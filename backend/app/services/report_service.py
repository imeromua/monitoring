import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text, bindparam

from app.config import settings

logger = logging.getLogger(__name__)

KYIV_TZ = ZoneInfo("Europe/Kyiv")

# ─── Кольорова палітра ──────────────────────────────────────────────────────
HEADER_BG      = "1A56A8"   # темно-синій заголовок
HEADER_FG      = "FFFFFF"   # білий текст заголовка
ROW_ODD_BG     = "EBF3FB"   # блакитний рядок (непарний)
ROW_EVEN_BG    = "FFFFFF"   # білий рядок (парний)
PROMO_BG       = "FFF3CD"   # жовтий — акційна ціна
PROMO_FG       = "7D5A00"   # коричневий текст акції
COMPETITOR_BG  = "FFE2E2"   # рожевий — новинка конкурента
COMPETITOR_FG  = "8B0000"   # темно-червоний текст
VARIANT_FG     = "5A5A8A"   # фіолетовий курсив — варіант
MISSING_BG     = "F0F0F0"   # сірий — відсутній товар
BORDER_COLOR   = "B8CCE4"   # колір меж

# ─── Тонка межа для всіх клітинок ──────────────────────────────────────────
_thin = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ─── SQL-запит ──────────────────────────────────────────────────────────────
# Singleton engine для синхронних операцій (Celery)
_sync_engine = None


def get_sync_engine():
    global _sync_engine
    if _sync_engine is None:
        sync_url = settings.DATABASE_URL.replace(
            "postgresql+asyncpg", "postgresql+psycopg2"
        )
        _sync_engine = create_engine(sync_url, pool_pre_ping=True)
    return _sync_engine


_REPORT_SQL = text("""
    SELECT
        mr.created_at AT TIME ZONE 'Europe/Kyiv'  AS "Дата/Час",
        u.full_name                                AS "Працівник",
        s.name                                     AS "Магазин",
        p.article_id                               AS "Артикул",
        COALESCE(mr.custom_name, p.name)           AS "Товар",
        mr.price                                   AS "Ціна",
        CASE WHEN mr.is_promo    THEN 'Акція'   ELSE 'Звичайна' END AS "Тип ціни",
        CASE WHEN mr.is_missing  THEN 'Так'     ELSE 'Ні'       END AS "Відсутній",
        mr.result_type                             AS "Тип запису"
    FROM monitoring_results mr
    JOIN monitoring_sessions ms ON mr.session_id = ms.id
    JOIN users u                ON ms.user_id    = u.id
    JOIN stores s               ON ms.store_id   = s.id
    LEFT JOIN products p        ON mr.product_id = p.id
    WHERE
        (:session_id IS NULL OR ms.id        = :session_id)
        AND (:store_id   IS NULL OR s.id     = :store_id)
        AND (:date_from  IS NULL OR mr.created_at >= :date_from ::date)
        AND (:date_to    IS NULL OR mr.created_at  < :date_to  ::date + interval '1 day')
    ORDER BY mr.created_at
""").bindparams(
    bindparam("session_id", value=None),
    bindparam("store_id",   value=None),
    bindparam("date_from",  value=None),
    bindparam("date_to",    value=None),
)


# ─── Стилізація Excel ────────────────────────────────────────────────────────

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _apply_header(ws, n_cols: int):
    """Стилізує рядок-заголовок."""
    hdr_fill = _make_fill(HEADER_BG)
    hdr_font = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = CELL_BORDER

    ws.row_dimensions[1].height = 30


def _style_data_rows(ws, df: pd.DataFrame):
    """Стилізує рядки даних відповідно до типу запису."""
    result_col   = df.columns.get_loc("Тип запису")  + 1
    promo_col    = df.columns.get_loc("Тип ціни")    + 1
    missing_col  = df.columns.get_loc("Відсутній")   + 1
    name_col     = df.columns.get_loc("Товар")        + 1
    price_col    = df.columns.get_loc("Ціна")         + 1
    n_cols       = len(df.columns)

    odd_fill        = _make_fill(ROW_ODD_BG)
    even_fill       = _make_fill(ROW_EVEN_BG)
    promo_fill      = _make_fill(PROMO_BG)
    competitor_fill = _make_fill(COMPETITOR_BG)
    missing_fill    = _make_fill(MISSING_BG)

    base_font    = Font(name="Calibri", size=10)
    promo_font   = Font(name="Calibri", size=10, bold=True, color=PROMO_FG)
    comp_font    = Font(name="Calibri", size=10, bold=True, color=COMPETITOR_FG)
    variant_font = Font(name="Calibri", size=10, italic=True, color=VARIANT_FG)
    missing_font = Font(name="Calibri", size=10, color="888888")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        result_type = str(row_data.get("Тип запису", ""))
        is_promo    = str(row_data.get("Тип ціни", "")) == "Акція"
        is_missing  = str(row_data.get("Відсутній", "")) == "Так"

        # Базовий фон рядка
        if result_type == "competitor_new":
            row_fill = competitor_fill
            row_font = comp_font
        elif is_missing:
            row_fill = missing_fill
            row_font = missing_font
        elif is_promo:
            row_fill = promo_fill
            row_font = promo_font
        elif row_idx % 2 == 0:
            row_fill = even_fill
            row_font = base_font
        else:
            row_fill = odd_fill
            row_font = base_font

        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill   = row_fill
            cell.border = CELL_BORDER
            cell.alignment = center_align if col != name_col else left_align

            # Спеціальний шрифт для назви варіанту
            if col == name_col and result_type == "variant":
                cell.font = variant_font
            else:
                cell.font = row_font

        # Числовий формат ціни
        price_cell = ws.cell(row=row_idx, column=price_col)
        if price_cell.value is not None:
            price_cell.number_format = '#,##0.00 ₴'

        ws.row_dimensions[row_idx].height = 18


def _auto_column_width(ws, df: pd.DataFrame):
    """Авто-ширина стовпців: по максимальному тексту + відступ."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)

        # Ширина заголовка
        max_len = len(str(col_name))
        # Ширина даних
        for val in df[col_name].astype(str):
            max_len = max(max_len, len(val))

        # Обмеження: не менше 10, не більше 50
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)


def _add_legend(ws, n_cols: int, n_rows: int):
    """Додає невеличку легенду під таблицею."""
    legend_row = n_rows + 3
    legends = [
        (COMPETITOR_BG, COMPETITOR_FG, "🔴 Новинка конкурента"),
        (PROMO_BG,      PROMO_FG,      "🟡 Акційна ціна"),
        (MISSING_BG,    "888888",      "⬜ Відсутній товар"),
        (ROW_ODD_BG,    VARIANT_FG,    "🔵 Варіант товару (курсив)"),
    ]
    for i, (bg, fg, label) in enumerate(legends):
        cell = ws.cell(row=legend_row + i, column=1)
        cell.value = label
        cell.fill  = _make_fill(bg)
        cell.font  = Font(name="Calibri", size=9, color=fg)
        cell.alignment = Alignment(horizontal="left")


def build_report_sync(
    session_id: Optional[int] = None,
    store_id:   Optional[int] = None,
    date_from:  Optional[date] = None,
    date_to:    Optional[date] = None,
) -> str:
    """
    Генерує красивий .xlsx звіт (синхронно, для Celery або прямого виклику).
    Повертає повний шлях до файлу.
    """
    engine = get_sync_engine()

    params = {
        "session_id": session_id,
        "store_id":   store_id,
        "date_from":  str(date_from) if date_from else None,
        "date_to":    str(date_to)   if date_to   else None,
    }

    with engine.connect() as conn:
        df = pd.read_sql(_REPORT_SQL, conn, params=params)

    # Форматуємо Дата/Час
    if "Дата/Час" in df.columns:
        df["Дата/Час"] = pd.to_datetime(df["Дата/Час"]).dt.strftime("%d.%m.%Y %H:%M")

    reports_path = settings.reports_path
    timestamp    = datetime.now(KYIV_TZ).strftime("%Y%m%d_%H%M%S")
    filename     = f"report_{session_id or 'custom'}_{timestamp}.xlsx"
    filepath     = reports_path / filename

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Звіт")
        ws = writer.sheets["Звіт"]

        n_cols = len(df.columns)
        n_rows = len(df)

        # Заморожуємо перший рядок (заголовок)
        ws.freeze_panes = "A2"

        # Стилізація
        _apply_header(ws, n_cols)
        _style_data_rows(ws, df)
        _auto_column_width(ws, df)
        _add_legend(ws, n_cols, n_rows)

        # Авто-фільтр
        ws.auto_filter.ref = ws.dimensions

        # Назва аркуша
        if session_id:
            ws.title = f"Сесія {session_id}"
        else:
            ws.title = "Звіт"

    logger.info("Report generated: %s (%d rows)", filepath, n_rows)
    return str(filepath)


# ─── Email-відправка ─────────────────────────────────────────────────────────

def send_report_email(
    filepath: str,
    recipients: list[str],
    session_id: int,
    extra_subject: str = "",
):
    """
    Відправляє Excel-звіт на пошту.
    Кидає виняток при помилці — Celery-задача перехопить і зробить retry.
    """
    filename = Path(filepath).name
    subject  = f"📊 Звіт моніторингу #{session_id}"
    if extra_subject:
        subject += f" — {extra_subject}"

    msg            = MIMEMultipart()
    msg["From"]    = settings.SMTP_USER
    msg["To"]      = ", ".join(recipients)
    msg["Subject"] = subject

    body = (
        f"Вітаємо!\n\n"
        f"У вкладенні — звіт цінового моніторингу по сесії #{session_id}.\n"
        f"Файл: {filename}\n\n"
        f"З повагою,\nСистема моніторингу Store Check"
    )
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{filename}"',
        )
        msg.attach(part)

    logger.info("Sending report email to %s", recipients)
    with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
        server.starttls()
        server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
        server.sendmail(settings.SMTP_USER, recipients, msg.as_string())
    logger.info("Report email sent successfully")
